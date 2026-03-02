"""Excel roster parsing and filtering."""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class RosterError(Exception):
    """Roster parsing error."""
    pass


def load_roster(excel_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """
    Load member roster from Excel file.
    
    Supports both old and new column names:
    - member_id / Badge ID
    - full_name / Full Name (Vietnamese)
    - email / Email
    - birthday / Date Of Birth
    - photo_path / Image URL
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of worksheet (auto-detects if None)
        
    Returns:
        List of member dictionaries
        
    Raises:
        RosterError: If file is invalid or required columns missing
    """
    excel_file = Path(excel_path)
    
    if not excel_file.exists():
        raise RosterError(f"Excel file not found: {excel_path}")
    
    # Auto-detect sheet name if not provided
    if sheet_name is None:
        try:
            wb = load_workbook(excel_path)
            # Try common sheet names
            for name in ['Members', 'Sheet1', 'Employees', 'Staff']:
                if name in wb.sheetnames:
                    sheet_name = name
                    break
            
            if sheet_name is None:
                # Use the first sheet
                sheet_name = wb.sheetnames[0]
                logger.info(f"Using sheet: {sheet_name}")
        except Exception as e:
            raise RosterError(f"Failed to detect sheet name: {e}")
    
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except Exception as e:
        raise RosterError(f"Failed to read Excel file: {e}")
    
    # Map new column names to old names for compatibility
    column_mapping = {
        'Badge ID': 'member_id',
        'Full Name (Vietnamese)': 'full_name',
        'Email': 'email',
        'Date Of Birth': 'birthday',
        'Image URL': 'photo_path'
    }
    
    # Rename columns if new names are found
    for new_name, old_name in column_mapping.items():
        if new_name in df.columns and old_name not in df.columns:
            df = df.rename(columns={new_name: old_name})
    
    # Validate required columns
    required_columns = ['member_id', 'full_name', 'email', 'birthday', 'photo_path']
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise RosterError(f"Missing required columns: {missing}")
    
    # Convert to list of dicts
    members = []
    for idx, row in df.iterrows():
        member = {}
        
        # Required fields
        member['member_id'] = str(row['member_id']).strip()
        member['full_name'] = str(row['full_name']).strip()
        member['email'] = str(row['email']).strip()
        member['photo_path'] = str(row['photo_path']).strip()
        
        # Birthday - handle both Excel dates and ISO strings
        birthday = row['birthday']
        if pd.isna(birthday):
            logger.warning(f"Member {member['member_id']} has missing birthday")
            continue
        
        if isinstance(birthday, datetime):
            member['birthday'] = birthday
        elif isinstance(birthday, str):
            # Try multiple date formats
            date_formats = [
                '%Y-%m-%d',  # 2000-06-20
                '%Y/%m/%d',  # 2000/06/20
                '%d/%m/%Y',  # 20/06/2000
                '%d-%m-%Y',  # 20-06-2000
                '%b %d, %Y',  # Jun 20, 2000
                '%B %d, %Y',  # June 20, 2000
                '%d %b %Y',   # 20 Jun 2000
                '%d %B %Y',   # 20 June 2000
                '%m/%d/%Y',   # 06/20/2000
            ]
            
            parsed_date = None
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(birthday.strip(), fmt)
                    break
                except ValueError:
                    continue
            
            if parsed_date:
                member['birthday'] = parsed_date
            else:
                logger.warning(f"Member {member['member_id']} has invalid birthday format: {birthday}")
                continue
        else:
            logger.warning(f"Member {member['member_id']} has invalid birthday type: {type(birthday)}")
            continue
        
        # Optional fields
        member['enabled'] = True
        if 'enabled' in df.columns:
            enabled_val = row['enabled']
            if pd.notna(enabled_val):
                if isinstance(enabled_val, bool):
                    member['enabled'] = enabled_val
                elif isinstance(enabled_val, str):
                    member['enabled'] = enabled_val.lower() in ('true', 'yes', '1')
                else:
                    member['enabled'] = bool(enabled_val)
        
        member['team'] = ''
        if 'team' in df.columns and pd.notna(row['team']):
            member['team'] = str(row['team']).strip()
        
        member['greeting_message'] = ''
        if 'greeting_message' in df.columns and pd.notna(row['greeting_message']):
            member['greeting_message'] = str(row['greeting_message']).strip()
        
        # Add any extra columns as-is
        for col in df.columns:
            if col not in member:
                val = row[col]
                if pd.notna(val):
                    member[col] = val
        
        members.append(member)
    
    logger.info(f"Loaded {len(members)} members from roster")
    return members


def filter_birthdays_today(
    members: List[Dict[str, Any]],
    today: Optional[datetime] = None
) -> List[Dict[str, Any]]:
    """
    Filter members with birthdays today.
    
    Args:
        members: List of member dictionaries
        today: Date to check (default: current date)
        
    Returns:
        List of members with birthdays today
    """
    if today is None:
        today = datetime.now()
    
    birthdays_today = []
    
    for member in members:
        if not member.get('enabled', True):
            continue
        
        birthday = member.get('birthday')
        if not isinstance(birthday, datetime):
            continue
        
        # Compare month and day only
        if birthday.month == today.month and birthday.day == today.day:
            birthdays_today.append(member)
    
    logger.info(f"Found {len(birthdays_today)} birthday(s) today")
    return birthdays_today


def filter_birthdays_on_date(
    members: List[Dict[str, Any]],
    target_date: datetime
) -> List[Dict[str, Any]]:
    """
    Filter members with birthdays on specific date.
    
    Args:
        members: List of member dictionaries
        target_date: Target date to check
        
    Returns:
        List of members with birthdays on that date
    """
    return filter_birthdays_today(members, target_date)


def validate_member_photos(
    members: List[Dict[str, Any]],
    base_path: Optional[Path] = None
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Validate that photo files exist.
    
    Args:
        members: List of member dictionaries
        base_path: Base path for relative photo paths (default: current directory)
        
    Returns:
        Tuple of (valid_members, invalid_photo_paths)
    """
    if base_path is None:
        base_path = Path.cwd()
    
    valid = []
    invalid = []
    
    for member in members:
        photo_path_str = member.get('photo_path', '')
        
        if not photo_path_str:
            invalid.append(f"{member['member_id']}: empty photo path")
            continue
        
        # Check if it's a URL (can be constructed later with base_url from config)
        if photo_path_str.startswith('http://') or photo_path_str.startswith('https://'):
            # It's a full URL, consider it valid
            member['_resolved_photo_path'] = photo_path_str
            valid.append(member)
            continue
        
        photo_path = Path(photo_path_str)
        
        # Try as absolute path first, then relative to base_path
        if not photo_path.is_absolute():
            photo_path = base_path / photo_path
        
        if photo_path.exists():
            member['_resolved_photo_path'] = str(photo_path.resolve())
            valid.append(member)
        else:
            # Photo path might be a relative URL path to be prefixed with base_url later
            # Don't reject it automatically - let the renderer handle it
            member['_resolved_photo_path'] = photo_path_str
            valid.append(member)
    
    if invalid:
        logger.warning(f"Invalid photo paths:\n  " + "\n  ".join(invalid))
    
    return valid, invalid
